from datetime import datetime, timezone
from io import BytesIO

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.parsing.models import Participant

def build_excel(
    participants: list[Participant],
    mentions: list[str],
) -> BytesIO:
    wb = Workbook()

    ws = wb.active
    ws.title = "Participants"

    export_date = datetime.now(timezone.utc).date().isoformat()

    headers = ["export_date", "username", "full_name", "bio/about"]
    ws.append(headers)

    for p in participants:
        ws.append([
            export_date,
            p.username or "",
            p.full_name or "",
            p.bio or "",
        ])

    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        max_len = 0
        for cell in ws[letter]:
            v = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(v))
        ws.column_dimensions[letter].width = min(max_len + 2, 50)

    ws2 = wb.create_sheet("Mentions")
    ws2.append(["username_mentioned"])
    for m in mentions:
        ws2.append([m])

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out
